############## Mycelium Version 0.18.11 of 2024.02.16 ##############
#Documents routines are deprecated. All the functions moved to Processors and became DOCXLoader and XLSXLoader.

from Mycelium import UnifiedPrompt

from docx import Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import Table
from docx.text.paragraph import Paragraph

import io
from PIL import Image

import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom

class DocxToPromptsConverter:
    def __init__(self, convert_urls=False):
        self.convert_urls = convert_urls

    def is_url(self, text):
        # Simple URL check - may require more sophisticated validation
        return text.startswith('http://') or text.startswith('https://')

    def process_paragraph(self, paragraph):
        prompts = []
        for run in paragraph.runs:
            text = run.text.strip()
            if text:
                if self.convert_urls and self.is_url(text):
                    prompts.append(UnifiedPrompt("url", text, "text/plain"))
                else:
                    if self.contains_chars_or_nums(text):
                        prompts.append(UnifiedPrompt("text", text, "text/plain"))
        return prompts

    def process_table_cell(self, cell, table_count, row_index, cell_index):
        cell_id = f"TableCell_{table_count}_{row_index}_{cell_index}"
        cell_prompts = [UnifiedPrompt("text", cell_id, "text/plain")]

        for paragraph in cell.paragraphs:
            cell_prompts.extend(self.process_paragraph(paragraph))
        
        return cell_prompts

    def process_table(self, table, table_count):
        table_prompts = []
        for row_index, row in enumerate(table.rows):
            for cell_index, cell in enumerate(row.cells):
                table_prompts.extend(self.process_table_cell(cell, table_count, row_index, cell_index))
        return table_prompts
    
    def contains_chars_or_nums(self, s):
        return any(char.isalpha() or char.isdigit() for char in s)

    def process_image(self, run):
        inline_shapes = run.element.xpath('.//a:blip')
        if inline_shapes:
            blip_element = inline_shapes[0]
            image_rid = blip_element.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
            image_part = run.part.related_parts[image_rid]
            image_bytes = io.BytesIO(image_part.blob)
            image = Image.open(image_bytes)
            return UnifiedPrompt("image", image, "image/jpeg")
        return None

    def convert(self, docx_file):
        doc = Document(docx_file)
        prompts = [UnifiedPrompt("text", "This content is extraxted from DOCX file.", "text/plain")]
        table_count = 0

        for element in doc.element.body:
            if isinstance(element, CT_P):
                paragraph = Paragraph(element, doc)
                for run in paragraph.runs:
                    if run.element.xpath('.//a:blip'):
                        image_prompt = self.process_image(run)
                        if image_prompt:
                            prompts.append(image_prompt)
                    else:
                        prompts.extend(self.process_paragraph(paragraph))
                        break  # Exit after processing the first non-image run
            elif isinstance(element, CT_Tbl):
                table = Table(element, doc)
                table_count += 1
                table_prompts = self.process_table(table, table_count)
                prompts.extend(table_prompts)

        return prompts

    @staticmethod
    def iter_block_items(parent):
        """
        Generate a reference to each paragraph and table child within parent,
        in document order. Each returned value is an instance of either Table or CT_P.
        """
        if isinstance(parent, Document):
            parent_elm = parent.element.body
        else:
            raise ValueError("Parent must be a Document object.")

        for child in parent_elm.iterchildren():
            if isinstance(child, CT_P):
                yield child
            elif isinstance(child, CT_Tbl):
                yield Table(child, parent)

    def print_prompts(self, prompts):
        for prompt in prompts:
            if prompt.content_type == "image":
                print("image")
            else:
                print(f"{prompt.content_type}: {prompt.content}")
                             
class XlsxToPromptsConverter:
    def convert(self, xlsx_input):
        if isinstance(xlsx_input, str):
            workbook = openpyxl.load_workbook(xlsx_input, data_only=True)
        elif isinstance(xlsx_input, io.BytesIO):
            workbook = openpyxl.load_workbook(filename=xlsx_input, data_only=True)
        else:
            raise ValueError("Input must be a file path or a BytesIO object")

        prompts = []

        for sheet in workbook.sheetnames:
            sheet_element = ET.Element("Sheet", name=sheet)
            worksheet = workbook[sheet]

            merged_cells_ranges = worksheet.merged_cells.ranges

            for row in worksheet.iter_rows():
                row_element = ET.SubElement(sheet_element, "Row")
                for cell in row:
                    # Check if cell is part of a merged cell and get the value
                    merged_cell_value = self.get_merged_cell_value(cell, merged_cells_ranges)
                    if merged_cell_value is not None:
                        cell_value = merged_cell_value
                    else:
                        cell_value = str(cell.value) if cell.value is not None else ''

                    if isinstance(cell, openpyxl.cell.cell.MergedCell):  # Check if it's a MergedCell
                        continue  # Skip processing for non-top-left cells in merged regions

                    cell_element = ET.SubElement(row_element, "Cell", column=cell.column_letter)
                    cell_element.text = cell_value

            xml_str = ET.tostring(sheet_element, encoding='utf-8').decode('utf-8')
            parsed_xml = minidom.parseString(xml_str)
            pretty_xml_str = parsed_xml.toprettyxml(indent="  ")

            prompts.append(UnifiedPrompt("text", pretty_xml_str, "text/plain"))
        return prompts

    def get_merged_cell_value(self, cell, merged_ranges):
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                top_left_cell = merged_range.start_cell
                return str(top_left_cell.value) if top_left_cell.value is not None else ''
        return None

    def print_prompts(self, prompts):
        for prompt in prompts:
            if prompt.content_type == "image":
                print("image")
            else:
                print(f"{prompt.content_type}: {prompt.content}")