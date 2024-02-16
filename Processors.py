############## Mycelium Version 0.18.11 of 2024.02.16 ##############

from Mycelium import Dialog, Message, UnifiedPrompt
import aiohttp
from bs4 import BeautifulSoup
from datetime import datetime
from docx import Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import Table
from docx.text.paragraph import Paragraph
import io
import json
import openpyxl
import os
from PIL import Image
import re
import requests
import uuid
import xml.etree.ElementTree as ET
from xml.dom import minidom

class MessageSplitter:
    def __init__ (self, acceptedRoles=[]):
        self.acceptedRoles = acceptedRoles
        
    def __rrshift__(self, other):
        if not isinstance(other, Dialog) and not isinstance(other, Dialog):
            raise TypeError("Message splitter is only aplicable to Dialog objects or lists of Dialog objects")
        if isinstance(other, Dialog):
            return self.__Process(other)
        if isinstance(other, list):
            result = []
            for dialog in other:
                result.append(self.__Process(dialog))
            return result
        
    def __Process(self, dialog):
        if isinstance(dialog, Dialog):
            result = []
            for msg in dialog.messages:
                if self.acceptedRoles == [] or msg.role in self.acceptedRoles:
                    result.append(Dialog(messages=[msg]))
            return result if len(result) > 1 else result[0]
        else:
            raise TypeError("Message splitter is only aplicable to Dialog objects or lists of Dialog objects")
        
class TextLineSplitter:
    def __init__ (self, lastMessageCount = 1, acceptedRoles=[]):
        self.lastMessageCount = lastMessageCount
        self.acceptedRoles = acceptedRoles
        
    def __rrshift__(self, other):
        if not isinstance(other, Dialog) and not isinstance(other, Dialog):
            raise TypeError("Message splitter is only aplicable to Dialog objects or lists of Dialog objects")
        if isinstance(other, Dialog):
            return self.__Process(other)
        if isinstance(other, list):
            result = []
            for dialog in other:
                result.append(self.__Process(dialog))
            return result
        
    def __Process(self, dialog):
        if isinstance(dialog, Dialog):
            i = 0
            result = []
            for msg in reversed(dialog.messages):
                if i >= self.lastMessageCount:
                    break
                if self.acceptedRoles == [] or msg.role in self.acceptedRoles:
                    for prompt in msg.unified_prompts:
                        if prompt.content_type == "text":
                            lines = prompt.content.splitlines()
                            for line in lines:
                                unified_prompt = UnifiedPrompt(content_type="text", content=line, mime_type="text/plain")
                                message = Message(unified_prompts=[unified_prompt], role = msg.role, sender_info=msg.sender_info, send_datetime=msg.send_datetime, diagnosticData=msg.diagnosticData, agentConfig=msg.agentConfig, billingData=msg.billingData, routingStrategy=msg.routingStrategy)
                                result.append(Dialog(messages=[message]))
                i += 1
            return result
        else:
            raise TypeError("Message splitter is only aplicable to Dialog objects")
        
class TextListSplitter:
    def __init__(self, lastMessageCount=1, acceptedRoles=[], removeListMarks=True):
        self.lastMessageCount = lastMessageCount
        self.acceptedRoles = acceptedRoles
        self.removeListMarks = removeListMarks

        # Extend the pattern to match more complex list bullets
        self.pattern = re.compile(r'^((\d+|[a-zA-Z])(\.\d+)*(\.[a-zA-Z])?[\.\)]\s|[-\*+]|--\s|–\s|—\s)')

    def __rrshift__(self, other):
        if not isinstance(other, Dialog) and not isinstance(other, Dialog):
            raise TypeError("Message splitter is only aplicable to Dialog objects or lists of Dialog objects")
        if isinstance(other, Dialog):
            return self.__Process(other)
        if isinstance(other, list):
            result = []
            for dialog in other:
                result.append(self.__Process(dialog))
            return result

    def __Process(self, dialog):
        if isinstance(dialog, Dialog):
            i = 0
            result = []
            for msg in reversed(dialog.messages):
                if i >= self.lastMessageCount:
                    break
                if self.acceptedRoles == [] or msg.role in self.acceptedRoles:
                    for prompt in msg.unified_prompts:
                        if prompt.content_type == "text":
                            lines = prompt.content.splitlines()
                            for line in lines:
                                # Strip leading spaces and tabs before checking for list marks
                                stripped_line = line.lstrip()
                                match = self.pattern.match(stripped_line)
                                if match:
                                    # Calculate content start index, adjusting for stripped leading whitespace
                                    content_start_index = match.end() + (len(line) - len(stripped_line))
                                    line_content = line[content_start_index:] if self.removeListMarks else line
                                    if len(line_content)>0 and len(re.sub(r"[ \t]", "", line_content)) > 0:
                                        unified_prompt = UnifiedPrompt(content_type="text", content=line_content, mime_type="text/plain")
                                        message = Message(unified_prompts=[unified_prompt], role=msg.role, sender_info=msg.sender_info, send_datetime=msg.send_datetime, diagnosticData=msg.diagnosticData, agentConfig=msg.agentConfig, billingData=msg.billingData, routingStrategy=msg.routingStrategy)
                                        result.append(Dialog(messages=[message]))
                i += 1
            return result
        else:
            raise TypeError("Message splitter is only applicable to Dialog objects")
        
class TextRegExpSplitter:
    '''
    Pattern Customization: The __init__ method now accepts a pattern parameter, allowing the user to specify any regular expression pattern for splitting the text.
    Remove Pattern Option: The removePattern parameter determines whether to remove the matched pattern from the beginning of each line in the content.
    Type Checks and Processing Logic: The processing logic remains largely the same, but it's now based on the custom regular expression pattern provided at initialization.
    '''
    def __init__(self, pattern, lastMessageCount=1, acceptedRoles=[], removePattern=True):
        self.lastMessageCount = lastMessageCount
        self.acceptedRoles = acceptedRoles
        self.removePattern = removePattern
        self.pattern = re.compile(pattern)

    def __rrshift__(self, other):
        if not isinstance(other, Dialog) and not isinstance(other, list):
            raise TypeError("Message splitter is only applicable to Dialog objects or lists of Dialog objects")
        if isinstance(other, Dialog):
            return self.__Process(other)
        elif isinstance(other, list):
            return [self.__Process(dialog) for dialog in other]

    def __Process(self, dialog):
        if not isinstance(dialog, Dialog):
            raise TypeError("Message splitter is only applicable to Dialog objects")
        
        result = []
        for msg in reversed(dialog.messages[:self.lastMessageCount]):
            if self.acceptedRoles == [] or msg.role in self.acceptedRoles:
                for prompt in msg.unified_prompts:
                    if prompt.content_type == "text":
                        lines = prompt.content.splitlines()
                        for line in lines:
                            stripped_line = line.lstrip()
                            match = self.pattern.match(stripped_line)
                            if match:
                                content_start_index = match.end() + (len(line) - len(stripped_line))
                                line_content = line[content_start_index:] if self.removePattern else line
                                if line_content.strip():  # Ensure the line has content besides whitespace
                                    unified_prompt = UnifiedPrompt(content_type="text", content=line_content, mime_type="text/plain")
                                    new_message = Message(unified_prompts=[unified_prompt], role=msg.role, sender_info=msg.sender_info, send_datetime=msg.send_datetime, diagnosticData=msg.diagnosticData, agentConfig=msg.agentConfig, billingData=msg.billingData, routingStrategy=msg.routingStrategy)
                                    result.append(Dialog(messages=[new_message]))
        return result

class DocxLoader:
    def __init__(self, docxFile, convert_urls=False):
        self.convert_urls = convert_urls
        self.docxFile = docxFile

    def __rshift__(self, other):
        if isinstance(other, Dialog):
            prompts = self.convert(self.docxFile)
            message = Message(unified_prompts=prompts, role = "user", sender_info="ComradeAI user", send_datetime=datetime.now())
            other.messages.append(message)
            return other
        else:
            raise TypeError("DocxLoader can be only applied to a Dialog object")
        
    def is_url(self, text):
        # Simple URL check - may require more sophisticated validation
        return text.startswith('http://') or text.startswith('https://')

    def process_paragraph(self, paragraph):
        prompts = []
        for run in paragraph.runs:
            text = run.text.strip()
            if text and len(text)>0 and len(re.sub(r"[ \t]", "", text)) > 0:
                if self.convert_urls and self.is_url(text):
                    prompts.append(UnifiedPrompt("url", text, "text/plain"))
                else:
                    if self.contains_chars_or_nums(text):
                        prompts.append(UnifiedPrompt("text", text, "text/plain"))
        return prompts

    def process_table_cell(self, cell, table_count, row_index, cell_index):
        cell_id = f"TableCell_{table_count}_{row_index}_{cell_index}"
        cell_prompts = [UnifiedPrompt("text", cell_id, "text/plain")]

        for paragraph in cell.paragraphs:
            cell_prompts.extend(self.process_paragraph(paragraph))
        
        return cell_prompts

    def process_table(self, table, table_count):
        table_prompts = []
        for row_index, row in enumerate(table.rows):
            for cell_index, cell in enumerate(row.cells):
                table_prompts.extend(self.process_table_cell(cell, table_count, row_index, cell_index))
        return table_prompts
    
    def contains_chars_or_nums(self, s):
        return any(char.isalpha() or char.isdigit() for char in s)

    def process_image(self, run):
        inline_shapes = run.element.xpath('.//a:blip')
        if inline_shapes:
            blip_element = inline_shapes[0]
            image_rid = blip_element.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
            image_part = run.part.related_parts[image_rid]
            image_bytes = io.BytesIO(image_part.blob)
            image = Image.open(image_bytes)
            return UnifiedPrompt("image", image, "image/jpeg")
        return None

    def convert(self, docx_file):
        doc = Document(docx_file)
        prompts = [UnifiedPrompt("text", "This content is loaded from DOCX file.", "text/plain")]
        table_count = 0

        for element in doc.element.body:
            if isinstance(element, CT_P):
                paragraph = Paragraph(element, doc)
                for run in paragraph.runs:
                    if run.element.xpath('.//a:blip'):
                        image_prompt = self.process_image(run)
                        if image_prompt:
                            prompts.append(image_prompt)
                    else:
                        prompts.extend(self.process_paragraph(paragraph))
                        break  # Exit after processing the first non-image run
            elif isinstance(element, CT_Tbl):
                table = Table(element, doc)
                table_count += 1
                table_prompts = self.process_table(table, table_count)
                prompts.extend(table_prompts)

        return prompts
    
class XlsxLoader:
    def __init__(self, xlsxFile = None, xlsxBytesArray = None, convert_urls=False):
        self.convert_urls = convert_urls
        self.xlsxFile = xlsxFile
        self.xlsxBytesArray = xlsxBytesArray
        if xlsxFile is None and xlsxBytesArray is None:
            raise ValueError ("Either provide a path to XLSX file as xlsxFile or xlsxBytesArray as xlsxBytesArray")

    def __rshift__(self, other):
        if isinstance(other, Dialog):
            xlsx = None
            if self.xlsxBytesArray is not None:
                xlsx = self.xlsxBytesArray
            elif self.xlsxFile is not None:
                xlsx = self.xlsxFile
            prompts = self.convert(xlsx)
            message = Message(unified_prompts=prompts, role ="user", sender_info="ComradeAI user", send_datetime=datetime.now())
            other.messages.append(message)
            return other
        else:
            raise TypeError("XlsxLoader can be only applied to a Dialog object")
        
    def convert(self, xlsx_input):
        if isinstance(xlsx_input, str):
            workbook = openpyxl.load_workbook(xlsx_input, data_only=True)
        elif isinstance(xlsx_input, io.BytesIO):
            workbook = openpyxl.load_workbook(filename=xlsx_input, data_only=True)
        else:
            raise ValueError("Input must be a file path or a BytesIO object")

        prompts = []

        for sheet in workbook.sheetnames:
            sheet_element = ET.Element("Sheet", name=sheet)
            worksheet = workbook[sheet]

            merged_cells_ranges = worksheet.merged_cells.ranges

            for row in worksheet.iter_rows():
                row_element = ET.SubElement(sheet_element, "Row")
                for cell in row:
                    # Check if cell is part of a merged cell and get the value
                    merged_cell_value = self.get_merged_cell_value(cell, merged_cells_ranges)
                    if merged_cell_value is not None:
                        cell_value = merged_cell_value
                    else:
                        cell_value = str(cell.value) if cell.value is not None else ''

                    if isinstance(cell, openpyxl.cell.cell.MergedCell):  # Check if it's a MergedCell
                        continue  # Skip processing for non-top-left cells in merged regions

                    cell_element = ET.SubElement(row_element, "Cell", column=cell.column_letter)
                    cell_element.text = cell_value

            xml_str = ET.tostring(sheet_element, encoding='utf-8').decode('utf-8')
            parsed_xml = minidom.parseString(xml_str)
            pretty_xml_str = parsed_xml.toprettyxml(indent="  ")

            prompts.append(UnifiedPrompt("text", pretty_xml_str, "text/plain"))
        return prompts

    def get_merged_cell_value(self, cell, merged_ranges):
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                top_left_cell = merged_range.start_cell
                return str(top_left_cell.value) if top_left_cell.value is not None else ''
        return None

    def print_prompts(self, prompts):
        for prompt in prompts:
            if prompt.content_type == "image":
                print("image")
            else:
                print(f"{prompt.content_type}: {prompt.content}")
                
class UrlLoader:
    def __init__(self, urlContainingContent) -> None:
        self.urlContainingContent = urlContainingContent
        
    def __rshift__(self, other):
        if isinstance(other, Dialog):
            links = self.extract_and_clean_links(self.urlContainingContent)
            prompts = []
            for link in links:
                pageContent = self.url_text(link)
                prompts.append(UnifiedPrompt(content_type="text", content=pageContent, mime_type="text/plain"))
            if len(prompts) > 0:
                message = Message(unified_prompts=prompts, role ="user", sender_info="ComradeAI user", send_datetime=datetime.now())
                other.messages.append(message)
            return other
        else:
            raise TypeError("Url loader can be only applied to a Dialog object")
        
    def extract_and_clean_links(self, text):
        pattern = r'\b(?:https?://)?(?:www\.)?(\S+\.\S+)'
        raw_links = re.findall(pattern, text)

        cleaned_links = []
        for link in raw_links:
            cleaned_link = re.sub(r'[.,/]+$', '', link)
            if not cleaned_link.startswith(('http://', 'https://')):
                cleaned_link = 'https://' + cleaned_link
            print(cleaned_link)
            if self.is_webpage(cleaned_link):
                cleaned_links.append(cleaned_link)
        return cleaned_links
          
    def is_webpage(self, url):
        try:
            response = requests.get(url)
            content_type = response.headers.get('Content-Type', '').lower()
            return content_type.startswith('text/html')
        except Exception as e:
            print(f"Error while checking the URL: {str(e)}")
            return False
        
    def url_text(self, msg):
        try:
            if msg[0] != "h":
                msg = msg[1:-1]

            headers = {
                'X-Requested-With': 'XMLHttpRequest',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:70.0) Gecko/20100101 Firefox/70.0'
            }

            response = requests.get(msg, headers=headers)
            html = response.content
            soup = BeautifulSoup(html, features="html.parser")
            for script in soup(["script", "style"]):
                script.extract()  # Remove these two elements from the soup.
            text = soup.get_text()
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = '\n'.join(chunk for chunk in chunks if chunk)
            return text

        except Exception as e:
            print(f"Failed to open link: {str(e)}")
            return None

class XlsxSplitter:
    def __init__ (self, splitMethod='row', lastMessageCount = 1, acceptedRoles=[]):
        self.splitMethod = splitMethod
        self.lastMessageCount = lastMessageCount
        self.acceptedRoles = acceptedRoles

    def split_xml_content(self, xml_string, splitMethod):
        root = ET.fromstring(xml_string)
        result = []
        for row in root.findall('.//Row'):
            if splitMethod == 'row':
                # For 'row' split method, concatenate Cell texts with a tab separator
                row_texts = [cell.text or '' for cell in row.findall('.//Cell')]
                result.append('\t'.join(row_texts))
            elif splitMethod == 'cell':
                # For 'cell' split method, add each Cell's text to the result list
                for cell in row.findall('.//Cell'):
                    result.append(cell.text or '')
        return result
    
    def __rrshift__(self, other):
        if not isinstance(other, Dialog) and not isinstance(other, Dialog):
            raise TypeError("Message splitter is only aplicable to Dialog objects or lists of Dialog objects")
        if isinstance(other, Dialog):
            return self.__Process(other)
        if isinstance(other, list):
            result = []
            for dialog in other:
                result.append(self.__Process(dialog))
            return result
        
    def __Process(self, dialog):
        if isinstance(dialog, Dialog):
            i = 0
            result = []
            for msg in reversed(dialog.messages):
                if i >= self.lastMessageCount:
                    break
                if self.acceptedRoles == [] or msg.role in self.acceptedRoles:
                    for prompt in msg.unified_prompts:
                        if prompt.content_type == "text":
                            lines = self.split_xml_content(xml_string = prompt.content, splitMethod = self.splitMethod)
                            for line in lines:
                                if len(line)>0 and len(re.sub(r"[ \t]", "", line)) > 0:
                                    unified_prompt = UnifiedPrompt(content_type="text", content=line, mime_type="text/plain")
                                    message = Message(unified_prompts=[unified_prompt], role = msg.role, sender_info=msg.sender_info, send_datetime=msg.send_datetime, diagnosticData=msg.diagnosticData, agentConfig=msg.agentConfig, billingData=msg.billingData, routingStrategy=msg.routingStrategy)
                                    result.append(Dialog(messages=[message]))
                i += 1
            return result
        else:
            raise TypeError("XML splitter is only aplicable to Dialog objects")
      
class DialogToFileDownloader:
    def __init__(self, dirPath = None):
        self.dirPath = dirPath if dirPath else os.getcwd() + "/downloads/"
        
    def __rrshift__(self, other):
        return self.__Process(other)
    
    def __Process(self, other):
        errorMessage = "Can only save content of Dialog object or [Dialog object]"
        if not isinstance(other, Dialog) and not isinstance(other, list):
            raise ValueError(errorMessage)
        if isinstance(other, Dialog):
            self.__Render(other)
        elif isinstance(other, list):
            for item in other:
                if not isinstance(item, Dialog):
                    raise ValueError(errorMessage)
                self.__Render(item)
        return other  
    
    def __Render(self, dialog):
        target_dir = os.path.join(self.dirPath, dialog.dialog_id)
        os.makedirs(target_dir, exist_ok=True)
        
        dialog_element = ET.Element("Dialog", id=dialog.dialog_id)
        for i, msg in enumerate(dialog.messages):
            msg_attr = {
                "id": str(i),
                "sender_info": msg.sender_info,
                "send_datetime": msg.send_datetime.strftime("%Y-%m-%d %H:%M:%S"),
            }
            message_element = ET.SubElement(dialog_element, "Message", msg_attr)
            
            billing_data_str = str(msg.billingData)  
            ET.SubElement(message_element, "BillingData", data=billing_data_str)
            
            for prompt in msg.unified_prompts:
                prompt_element = ET.SubElement(message_element, "Prompt", type=prompt.content_type)
                if prompt.content_type == "text":
                    prompt_element.text = prompt.content
                elif prompt.content_type == "image":
                    # Handle Pillow images
                    img_filename = uuid.uuid4().hex + ".png"  # Assuming PNG format; adjust if necessary
                    img_path = os.path.join(target_dir, img_filename)
                    prompt.content.save(img_path)  # Save the Pillow image
                    ET.SubElement(prompt_element, "Image", path=os.path.join(dialog.dialog_id, img_filename))
                elif prompt.content_type in ["audio", "video"]:
                    # Handle audio and video as bytes
                    file_extension = "mp3" if prompt.content_type == "audio" else "mp4"  # Simplified assumption; adjust based on actual mime_type or needs
                    filename = uuid.uuid4().hex + "." + file_extension
                    file_path = os.path.join(target_dir, filename)
                    with open(file_path, 'wb') as file:
                        file.write(prompt.content)
                    ET.SubElement(prompt_element, prompt.content_type.capitalize(), path=os.path.join(dialog.dialog_id, filename))
                elif prompt.content_type == "url":
                    # Handle URL
                    ET.SubElement(prompt_element, "URL", href=prompt.content)
                    
        # Write the XML and HTML files as before
        tree = ET.ElementTree(dialog_element)
        xml_path = os.path.join(target_dir, "content_manifest.xml")
        tree.write(xml_path, encoding='utf-8', xml_declaration=True)
        self.__GenerateHtml(target_dir, dialog_element)

    def __GenerateHtml(self, target_dir, dialog_element):
        html_path = os.path.join(target_dir, "content_view.html")
        with open(html_path, 'w', encoding='utf-8') as html_file:
            html_file.write('<!DOCTYPE html>\n<html>\n<head>\n<title>Dialog Content</title>\n</head>\n<body>\n')
            
            for message in dialog_element.findall('.//Message'):
                # Display message details
                sender_info = message.get('sender_info')
                send_datetime = message.get('send_datetime')
                html_file.write(f'<div><strong>Sender:</strong> {sender_info}, <strong>Time:</strong> {send_datetime}</div>\n')

                for prompt in message.findall('.//Prompt'):
                    content_type = prompt.attrib['type']
                    if content_type == "text":
                        html_file.write(f'<div>{prompt.text}</div>\n')
                    elif content_type == "image":
                        img_src = os.path.join(self.dirPath, prompt.find('Image').attrib['path']).replace(os.sep, '/')
                        html_file.write(f'<a href="{img_src}" target="_blank"><img src="{img_src}" alt="Image" style="width: 256px; height: 256px;"></a><br>\n')
                    elif content_type == "audio":
                        audio_src = os.path.join(self.dirPath, prompt.find('Audio').attrib['path']).replace(os.sep, '/')
                        html_file.write(f'<audio controls><source src="{audio_src}" type="audio/mpeg">Your browser does not support the audio element.</audio><br>\n')
                    elif content_type == "video":
                        video_src = os.path.join(self.dirPath, prompt.find('Video').attrib['path']).replace(os.sep, '/')
                        html_file.write(f'<video width="320" height="240" controls><source src="{video_src}" type="video/mp4">Your browser does not support the video tag.</video><br>\n')
                    elif content_type == "url":
                        url = prompt.find('URL').attrib['href']
                        html_file.write(f'<a href="{url}" target="_blank">{url}</a><br>\n')
            
            html_file.write('</body>\n</html>')
            
class DialogCollapser:
    def __init__(self, baseDialogNumber = 0):
        self.baseDialogNumber = baseDialogNumber
    
    def __rrshift__(self, other):
        return self.__Process(other)
    
    def __Process(self, other):
        errorMessage = "Can only collapce a list of Dialog objects"
        if isinstance(other, Dialog):
            return other
        elif isinstance(other, list):
            try:
                result = other[self.baseDialogNumber]
            except Exception as ex:
                print("BaseDialog error: " + str(ex))
            for i, dialog in enumerate(other):
                if isinstance(dialog, Dialog):
                    if i != self.baseDialogNumber:
                        result.messages.extend(dialog.messages)
                else:
                    raise TypeError(errorMessage)
            return result